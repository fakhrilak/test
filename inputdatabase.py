from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import mysql.connector
import random
import main
from openpyxl import Workbook
import pandas as pd
import xlsxwriter 
from tkinter import filedialog
import openpyxl as xl;
import os
from openpyxl import load_workbook
mydb = mysql.connector.connect(host="localhost",
                               user="fakhrilak",
                               passwd="fakhrilak",
                               database="wasit")
wb =  Workbook()
ws = wb.active
cursor = mydb.cursor()
root = Tk()
randAA=[]
c = 0
Apa = 0
Bpa = 0
Cpa = 0
Dpa = 0
Epa = 0
Fpa = 0
Gpa = 0
Hpa = 0
Api = 0
Bpi = 0
Cpi = 0
Dpi = 0
Epi = 0
Fpi = 0
Gpi = 0
Hpi = 0
total = str(' ')
def counter():
    global c
    print(c)
    c= int(c+1)
    if c == 1:
        button.config(text = "DATA PUTRA" )
    elif c == 2:
        button.config(text = "DATA PUTRI" )
    elif c == 3:
        button.config(text = "SUFFLE PUTRA" )
    elif c == 4:
        button.config(text = "SUFFLE PUTRI" )
    elif c == 5:
        button.config(text = "CETAK PUTRA" )
    elif c == 6:
        button.config(text = "CETAK PUTRI" )
    elif c > 6:
        button.config(text = "DATA PUTRA" )
        c = 1
        
def add():
    fno = ent.get()
    fnama = ent1.get()
    fkelas = ent2.get()
    fkontingen = ent3.get()
    fnama.rjust(10,"+")
    sql=("INSERT INTO kejuaraan VALUES(%s,%s,%s,%s)")
    cursor.execute(sql,(fno,fnama,fkelas,fkontingen))
    mydb.commit()
    print(fno)
    print(fnama)
    print(fkelas)
    print(fkontingen)
    ent.delete(0, END)
    ent1.delete(0, END)
    ent2.delete(0, END)
    ent3.delete(0, END)
    return True
def delet():
    fno = ent.get()
    h = "DELETE FROM kejuaraan WHERE no = %s "
    cursor.execute(h,(fno,))
    mydb.commit()
    print(fno,"Deleted")
    ent.delete(0, END)
def show():
    h = "DELETE FROM kejuaraan WHERE nama = ' '"
    cursor.execute(h)
    mydb.commit()
    sql = "SELECT * FROM kejuaraan ORDER BY NO DESC"
    cursor.execute(sql)
    rows = cursor.fetchall()
    total = cursor.rowcount
    print("Total data entries: "+ str(total))
    root=Tk()
    root.title("DAFTAR PESERTA")
    tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
    tv.pack(padx=20,pady=20)
    tv.heading(1, text="No")
    tv.heading(2, text="Nama")
    tv.heading(3, text="Kelas")
    tv.heading(4, text="Kontingen")
    for i in rows:
        tv.insert('','end', values=i)
def classA():
    global c,Apa,Api
    if c == 1:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Apa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Apa")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 2:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Api'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Api")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 3:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Apa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Apa = int(total)
        root=Tk()
        root.title("RANDOM KELAS Apa")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsApa.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" APa = ",Apa)
        main.cpyApa()
    elif c == 4:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Api'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Api = int(total)
        root=Tk()
        root.title("RANDOM KELAS Api")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsApi.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" APi = ",Api)
        main.cpyApi()
    elif c == 5:
        main.bgnApa()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsApa.xlsx")
    elif c == 6:
        main.bgnApi()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsApi.xlsx")
def classB():
    global c,Bpa,Bpi
    if c == 1:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Bpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Bpa")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 2:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Bpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Bpi")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 3:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Bpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Bpa = int(total)
        root=Tk()
        root.title("RANDOM KELAS Bpa")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsBpa.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" Bpa = ",Bpa)
        main.cpyBpa()
    elif c == 4:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Bpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Bpi = int(total)
        root=Tk()
        root.title("RANDOM KELAS Bpi")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsBpi.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" BPi = ",Bpi)
        main.cpyBpi()
    elif c == 5:
        main.bgnBpa()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsBpa.xlsx")
    elif c == 6:
        main.bgnBpi()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsBpi.xlsx")
def classC():
    global c,Cpa,Cpi
    if c == 1:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Cpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("RANDOM KELAS Cpa")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 2:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Cpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Cpi")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 3:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Cpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Cpa = int(total)
        root=Tk()
        root.title("RANDOM KELAS Cpa")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsCpa.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" Cpa = ",Cpa)
        main.cpyCpa()
    elif c == 4:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Cpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Cpi = int(total)
        root=Tk()
        root.title("RANDOM KELAS Cpi")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsCpi.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" CPi = ",Cpi)
        main.cpyCpi()
    elif c == 5:
        main.bgnCpa()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsCpa.xlsx")
    elif c == 6:
        main.bgnCpi()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsCpi.xlsx")
def classD():
    global c,Dpa,Dpi
    if c == 1:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Dpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR kelas DPA")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 2:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Dpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Dpi")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 3:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Dpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        print(rows)
        Dpa = int(total)
        root=Tk()
        root.title("RANDOM kelas DPA")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsDpa.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" Dpa = ",Dpa)
        main.cpyDpa()
    elif c == 4:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Dpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Dpi = int(total)
        root=Tk()
        root.title("RANDOM KELAS Dpi")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsDpi.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" Dpi = ",Dpi)
        main.cpyDpi()
    elif c == 5:
        main.bgnDpa()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsDpa.xlsx")
    elif c == 6:
        main.bgnDpi()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsDpi.xlsx")
def classE():
    global c,Epa,Epi
    if c == 1:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Epa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Epa")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 2:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Epi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Epi")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 3:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Epa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Epa = int(total)
        root=Tk()
        root.title("RANDOM KELAS Epa")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsEpa.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" Epa = ",Epa)
        main.cpyEpa()
    elif c == 4:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Epi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Epi = int(total)
        root=Tk()
        root.title("RANDOM KELAS Epi")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsEpi.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" EPi = ",Epi)
        main.cpyEpi()
    elif c == 5:
        main.bgnBpa()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsEpa.xlsx")
    elif c == 6:
        main.bgnBpi()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsEpi.xlsx")
def classF():
    global c,Fpa,Fpi
    if c == 1:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Fpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Fpa")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 2:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Fpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Fpi")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 3:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Fpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Fpa = int(total)
        root=Tk()
        root.title("RANDOM KELAS Fpa")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsFpa.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" Fpa = ",Fpa)
        main.cpyFpa()
    elif c == 4:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Fpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Fpi = int(total)
        root=Tk()
        root.title("RANDOM KELAS Fpi")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsFpi.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" Fpi = ",Fpi)
        main.cpyFpi()
    elif c == 5:
        main.bgnFpa()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsFpa.xlsx")
    elif c == 6:
        main.bgnFpi()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsFpi.xlsx")
def classG():
    global c,Gpa,Gpi
    if c == 1:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Gpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Gpa")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 2:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Gpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Gpi")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 3:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Gpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Gpa = int(total)
        root=Tk()
        root.title("RANDOM KELAS Gpa")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsGpa.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" Gpa = ",Gpa)
        main.cpyGpa()
    elif c == 4:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Gpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Gpi = int(total)
        root=Tk()
        root.title("RANDOM KELAS Gpi")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsGpi.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" Gpi = ",Gpi)
        main.cpyGpi()
    elif c == 5:
        main.bgnGpa()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsGpa.xlsx")
    elif c == 6:
        main.bgnGpi()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsGpi.xlsx")
def classH():
    global c,Hpa,Hpi
    if c == 1:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Hpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Hpa")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 2:
        cursor = mydb.cursor()
        h = "DELETE FROM kejuaraan WHERE nama = ' '"
        cursor.execute(h)
        mydb.commit()
        sql = "SELECT * FROM kejuaraan WHERE kelas='Hpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        print("Total data entries: "+ str(total))
        root=Tk()
        root.title("DAFTAR KELAS Hpi")
        tv = ttk.Treeview(root,columns=(1,2,3,4),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="No")
        tv.heading(2, text="Nama")
        tv.heading(3, text="Kelas")
        tv.heading(4, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
    elif c == 3:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Hpa'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Hpa = int(total)
        root=Tk()
        root.title("RANDOM KELAS Hpa")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsHpa.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" Hpa = ",Hpa)
        main.cpyHpa()
    elif c == 4:
        cursor = mydb.cursor()
        sql= "select nama,kontingen from kejuaraan WHERE kelas = 'Hpi'"
        cursor.execute(sql)
        rows = cursor.fetchall()
        total = cursor.rowcount
        random.shuffle(rows)
        randAA.append(rows)
        print("Total data entries: "+ str(total))
        Hpi = int(total)
        root=Tk()
        root.title("RANDOM KELAS Hpi")
        tv = ttk.Treeview(root,columns=(1,2),show= "headings",height="20")
        tv.pack(padx=20,pady=20)
        tv.heading(1, text="Nama")
        tv.heading(2, text="Kontingen")
        for i in rows:
            tv.insert('','end', values=i)
        df = pd.DataFrame(rows)
        writer = pd.ExcelWriter('/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsHpi.xlsx')
        df.to_excel(writer)
        writer.save()
        print(" Hpi = ",Hpi)
        main.cpyHpi()
    elif c == 5:
        main.bgnHpa()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/klsHpa.xlsx")
    elif c == 6:
        main.bgnHpi()
        os.remove("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRI/klsHpi.xlsx")
def livescore():
#########################################################################################livescore Apa
    c = partai.get()
    if c == 'APA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        print(a)
        print(b)
        c = ws.cell(row = 102, column = 2)
        c.value = a
        d = ws.cell(row = 103, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 104, column = 2)
        c.value = a
        d = ws.cell(row = 105, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 106, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 108, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 110, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 112, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 114, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 116, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 118, column = 2)
        c.value = a
        d = ws.cell(row = 113, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 120, column = 2)
        c.value = a
        d = ws.cell(row = 115, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 122, column = 2)
        c.value = a
        d = ws.cell(row = 117, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 124, column = 2)
        c.value = a
        d = ws.cell(row = 119, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 126, column = 2)
        c.value = a
        d = ws.cell(row = 121, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 128, column = 2)
        c.value = a
        d = ws.cell(row = 123, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 125, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 127, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 132, column = 2)
        c.value = a
        d = ws.cell(row = 129, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 134, column = 2)
        c.value = a
        d = ws.cell(row = 131, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 136, column = 2)
        c.value = a
        d = ws.cell(row = 133, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 138, column = 2)
        c.value = a
        d = ws.cell(row = 135, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 140, column = 2)
        c.value = a
        d = ws.cell(row = 137, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 142, column = 2)
        c.value = a
        d = ws.cell(row = 139, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 143, column = 2)
        c.value = a
        d = ws.cell(row = 141, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 145, column = 2)
        c.value = a
        d = ws.cell(row = 143, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'APA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 147, column = 2)
        c.value = a
        d = ws.cell(row = 145, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas APA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
######################################################################################livescore Bpa
    if c == 'BPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        print(a)
        print(b)
        c = ws.cell(row = 102, column = 2)
        c.value = a
        d = ws.cell(row = 103, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 104, column = 2)
        c.value = a
        d = ws.cell(row = 105, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 106, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 108, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 110, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 112, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 114, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 116, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 118, column = 2)
        c.value = a
        d = ws.cell(row = 113, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 120, column = 2)
        c.value = a
        d = ws.cell(row = 115, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 122, column = 2)
        c.value = a
        d = ws.cell(row = 117, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 124, column = 2)
        c.value = a
        d = ws.cell(row = 119, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 126, column = 2)
        c.value = a
        d = ws.cell(row = 121, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 128, column = 2)
        c.value = a
        d = ws.cell(row = 123, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 125, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 127, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 132, column = 2)
        c.value = a
        d = ws.cell(row = 129, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 134, column = 2)
        c.value = a
        d = ws.cell(row = 131, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 136, column = 2)
        c.value = a
        d = ws.cell(row = 133, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 138, column = 2)
        c.value = a
        d = ws.cell(row = 135, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 140, column = 2)
        c.value = a
        d = ws.cell(row = 137, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 142, column = 2)
        c.value = a
        d = ws.cell(row = 139, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 143, column = 2)
        c.value = a
        d = ws.cell(row = 141, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 145, column = 2)
        c.value = a
        d = ws.cell(row = 143, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'BPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 147, column = 2)
        c.value = a
        d = ws.cell(row = 145, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas BPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
##########################################################################################livescore CPA
    elif c == 'CPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        print(a)
        print(b)
        c = ws.cell(row = 102, column = 2)
        c.value = a
        d = ws.cell(row = 103, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 104, column = 2)
        c.value = a
        d = ws.cell(row = 105, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 106, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 108, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 110, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 112, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 114, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 116, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 118, column = 2)
        c.value = a
        d = ws.cell(row = 113, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 120, column = 2)
        c.value = a
        d = ws.cell(row = 115, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 122, column = 2)
        c.value = a
        d = ws.cell(row = 117, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 124, column = 2)
        c.value = a
        d = ws.cell(row = 119, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 126, column = 2)
        c.value = a
        d = ws.cell(row = 121, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 128, column = 2)
        c.value = a
        d = ws.cell(row = 123, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 125, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 127, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 132, column = 2)
        c.value = a
        d = ws.cell(row = 129, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 134, column = 2)
        c.value = a
        d = ws.cell(row = 131, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 136, column = 2)
        c.value = a
        d = ws.cell(row = 133, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 138, column = 2)
        c.value = a
        d = ws.cell(row = 135, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 140, column = 2)
        c.value = a
        d = ws.cell(row = 137, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 142, column = 2)
        c.value = a
        d = ws.cell(row = 139, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 143, column = 2)
        c.value = a
        d = ws.cell(row = 141, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 145, column = 2)
        c.value = a
        d = ws.cell(row = 143, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'CPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 147, column = 2)
        c.value = a
        d = ws.cell(row = 145, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas CPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
#########################################################################################livescore DPA
    elif c == 'DPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        print(a)
        print(b)
        c = ws.cell(row = 102, column = 2)
        c.value = a
        d = ws.cell(row = 103, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 104, column = 2)
        c.value = a
        d = ws.cell(row = 105, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 106, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 108, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 110, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 112, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 114, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 116, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 118, column = 2)
        c.value = a
        d = ws.cell(row = 113, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 120, column = 2)
        c.value = a
        d = ws.cell(row = 115, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 122, column = 2)
        c.value = a
        d = ws.cell(row = 117, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 124, column = 2)
        c.value = a
        d = ws.cell(row = 119, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 126, column = 2)
        c.value = a
        d = ws.cell(row = 121, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 128, column = 2)
        c.value = a
        d = ws.cell(row = 123, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 125, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 127, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 132, column = 2)
        c.value = a
        d = ws.cell(row = 129, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 134, column = 2)
        c.value = a
        d = ws.cell(row = 131, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 136, column = 2)
        c.value = a
        d = ws.cell(row = 133, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 138, column = 2)
        c.value = a
        d = ws.cell(row = 135, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 140, column = 2)
        c.value = a
        d = ws.cell(row = 137, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 142, column = 2)
        c.value = a
        d = ws.cell(row = 139, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 143, column = 2)
        c.value = a
        d = ws.cell(row = 141, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 145, column = 2)
        c.value = a
        d = ws.cell(row = 143, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'DPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 147, column = 2)
        c.value = a
        d = ws.cell(row = 145, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas DPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
#########################################################################################livescore EPA
    elif c == 'EPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        print(a)
        print(b)
        c = ws.cell(row = 102, column = 2)
        c.value = a
        d = ws.cell(row = 103, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 104, column = 2)
        c.value = a
        d = ws.cell(row = 105, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 106, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 108, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 110, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 112, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 114, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 116, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 118, column = 2)
        c.value = a
        d = ws.cell(row = 113, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 120, column = 2)
        c.value = a
        d = ws.cell(row = 115, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 122, column = 2)
        c.value = a
        d = ws.cell(row = 117, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 124, column = 2)
        c.value = a
        d = ws.cell(row = 119, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 126, column = 2)
        c.value = a
        d = ws.cell(row = 121, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 128, column = 2)
        c.value = a
        d = ws.cell(row = 123, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 125, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 127, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 132, column = 2)
        c.value = a
        d = ws.cell(row = 129, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 134, column = 2)
        c.value = a
        d = ws.cell(row = 131, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 136, column = 2)
        c.value = a
        d = ws.cell(row = 133, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 138, column = 2)
        c.value = a
        d = ws.cell(row = 135, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 140, column = 2)
        c.value = a
        d = ws.cell(row = 137, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 142, column = 2)
        c.value = a
        d = ws.cell(row = 139, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 143, column = 2)
        c.value = a
        d = ws.cell(row = 141, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 145, column = 2)
        c.value = a
        d = ws.cell(row = 143, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'EPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 147, column = 2)
        c.value = a
        d = ws.cell(row = 145, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas EPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END) 
#########################################################################################livescore FPA
    elif c == 'FPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        print(a)
        print(b)
        c = ws.cell(row = 102, column = 2)
        c.value = a
        d = ws.cell(row = 103, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 104, column = 2)
        c.value = a
        d = ws.cell(row = 105, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 106, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 108, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 110, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 112, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 114, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 116, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 118, column = 2)
        c.value = a
        d = ws.cell(row = 113, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 120, column = 2)
        c.value = a
        d = ws.cell(row = 115, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 122, column = 2)
        c.value = a
        d = ws.cell(row = 117, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 124, column = 2)
        c.value = a
        d = ws.cell(row = 119, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 126, column = 2)
        c.value = a
        d = ws.cell(row = 121, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 128, column = 2)
        c.value = a
        d = ws.cell(row = 123, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 125, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 127, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 132, column = 2)
        c.value = a
        d = ws.cell(row = 129, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 134, column = 2)
        c.value = a
        d = ws.cell(row = 131, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 136, column = 2)
        c.value = a
        d = ws.cell(row = 133, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 138, column = 2)
        c.value = a
        d = ws.cell(row = 135, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 140, column = 2)
        c.value = a
        d = ws.cell(row = 137, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 142, column = 2)
        c.value = a
        d = ws.cell(row = 139, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 143, column = 2)
        c.value = a
        d = ws.cell(row = 141, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 145, column = 2)
        c.value = a
        d = ws.cell(row = 143, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'FPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 147, column = 2)
        c.value = a
        d = ws.cell(row = 145, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas FPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
#########################################################################################livescore GPA
    elif c == 'GPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        print(a)
        print(b)
        c = ws.cell(row = 102, column = 2)
        c.value = a
        d = ws.cell(row = 103, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 104, column = 2)
        c.value = a
        d = ws.cell(row = 105, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 106, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 108, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 110, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 112, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 114, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 116, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 118, column = 2)
        c.value = a
        d = ws.cell(row = 113, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 120, column = 2)
        c.value = a
        d = ws.cell(row = 115, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 122, column = 2)
        c.value = a
        d = ws.cell(row = 117, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 124, column = 2)
        c.value = a
        d = ws.cell(row = 119, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 126, column = 2)
        c.value = a
        d = ws.cell(row = 121, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 128, column = 2)
        c.value = a
        d = ws.cell(row = 123, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 125, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 127, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 132, column = 2)
        c.value = a
        d = ws.cell(row = 129, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 134, column = 2)
        c.value = a
        d = ws.cell(row = 131, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 136, column = 2)
        c.value = a
        d = ws.cell(row = 133, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 138, column = 2)
        c.value = a
        d = ws.cell(row = 135, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 140, column = 2)
        c.value = a
        d = ws.cell(row = 137, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 142, column = 2)
        c.value = a
        d = ws.cell(row = 139, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 143, column = 2)
        c.value = a
        d = ws.cell(row = 141, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 145, column = 2)
        c.value = a
        d = ws.cell(row = 143, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'GPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 147, column = 2)
        c.value = a
        d = ws.cell(row = 145, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas GPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
#########################################################################################livescore HPA
    if c == 'HPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        print(a)
        print(b)
        c = ws.cell(row = 102, column = 2)
        c.value = a
        d = ws.cell(row = 103, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 104, column = 2)
        c.value = a
        d = ws.cell(row = 105, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 106, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 108, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 110, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 112, column = 2)
        c.value = a
        d = ws.cell(row = 107, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 114, column = 2)
        c.value = a
        d = ws.cell(row = 109, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 116, column = 2)
        c.value = a
        d = ws.cell(row = 111, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 118, column = 2)
        c.value = a
        d = ws.cell(row = 113, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 120, column = 2)
        c.value = a
        d = ws.cell(row = 115, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 122, column = 2)
        c.value = a
        d = ws.cell(row = 117, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 124, column = 2)
        c.value = a
        d = ws.cell(row = 119, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 126, column = 2)
        c.value = a
        d = ws.cell(row = 121, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 128, column = 2)
        c.value = a
        d = ws.cell(row = 123, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 125, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 130, column = 2)
        c.value = a
        d = ws.cell(row = 127, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 132, column = 2)
        c.value = a
        d = ws.cell(row = 129, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 134, column = 2)
        c.value = a
        d = ws.cell(row = 131, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 136, column = 2)
        c.value = a
        d = ws.cell(row = 133, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 138, column = 2)
        c.value = a
        d = ws.cell(row = 135, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 140, column = 2)
        c.value = a
        d = ws.cell(row = 137, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 142, column = 2)
        c.value = a
        d = ws.cell(row = 139, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 143, column = 2)
        c.value = a
        d = ws.cell(row = 141, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 145, column = 2)
        c.value = a
        d = ws.cell(row = 143, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
    elif c == 'HPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        ws = wb.get_sheet_by_name("Sheet1")
        a = kuning.get()
        b = biru.get()
        c = ws.cell(row = 147, column = 2)
        c.value = a
        d = ws.cell(row = 145, column = 2)
        d.value = b
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/kelas HPA.xlsx")
        partai.delete(0, END)
        kuning.delete(0, END)
        biru.delete(0, END)
#########################################################################################UPDATE JADWAL
def updatejadwal():
    c = read.get()
    d = 2 * int(write.get())
    B = 7
    E = 0
    F = 8
    A = 11 + d
    
############################################################################writeexcel to jadwal APA
#########################################################################
    if c == 'APA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA1"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B102"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B103"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C102"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D102"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C103"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D103"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA2"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B104"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B105"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C104"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D104"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C105"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D105"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA3"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B106"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B107"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C106"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D106"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C107"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D107"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA4"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B108"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B109"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C108"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D108"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C109"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D109"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA5"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B110"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B111"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C110"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C111"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D111"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA6"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B112"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B113"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C112"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D112"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C113"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D113"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA7"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B114"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B115"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C114"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D114"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C115"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D115"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA8"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B116"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B117"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C116"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D116"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C117"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D117"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA9"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B118"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B119"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C118"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D118"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C119"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D119"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA10"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B120"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B121"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C120"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C121"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D121"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA11"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B122"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B123"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C122"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D122"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C123"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D123"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA12"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B124"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B125"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C124"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D124"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C125"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D125"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA13"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B126"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B127"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C126"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D126"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C127"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D127"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA14"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B128"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B129"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C128"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D128"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C129"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D129"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA15"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B130"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B131"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C130"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D130"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C131"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D131"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA16"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B132"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B133"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C132"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D132"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C133"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D133"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA17"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B134"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B135"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C134"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D134"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C135"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D135"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA18"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B136"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B137"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C136"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D136"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C137"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D137"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA19"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B138"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B139"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C138"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D138"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C139"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D139"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA20"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B140"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B141"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C140"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D140"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C141"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D141"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA21"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B142"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B143"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C142"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D142"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C143"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D143"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA22"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B144"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B145"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C144"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D144"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C145"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D145"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA23"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B146"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B147"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C146"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D146"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C147"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D147"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA24"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B148"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B149"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C148"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D148"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C149"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D149"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'APA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "APA25"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Apa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B150"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!B151"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C150"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D150"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!C151"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas APA.xlsx]Sheet1'!D151"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
###################################################################################
    elif c == 'BPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA1"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B102"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B103"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C102"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D102"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C103"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D103"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA2"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B104"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B105"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C104"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D104"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C105"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D105"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA3"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B106"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B107"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C106"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D106"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C107"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D107"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA4"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B108"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B109"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C108"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D108"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C109"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D109"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA5"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B110"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B111"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C110"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C111"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D111"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA6"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B112"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B113"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C112"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D112"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C113"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D113"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA7"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B114"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B115"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C114"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D114"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C115"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D115"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA8"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B116"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B117"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C116"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D116"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C117"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D117"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA9"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B118"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B119"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C118"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D118"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C119"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D119"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA10"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B120"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B121"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C120"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C121"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D121"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA11"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B122"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B123"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C122"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D122"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C123"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D123"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA12"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B124"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B125"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C124"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D124"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C125"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D125"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA13"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B126"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B127"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C126"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D126"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C127"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D127"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA14"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B128"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B129"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C128"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D128"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C129"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D129"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA15"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B130"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B131"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C130"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D130"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C131"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D131"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA16"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B132"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B133"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C132"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D132"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C133"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D133"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA17"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B134"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B135"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C134"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D134"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C135"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D135"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA18"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B136"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B137"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C136"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D136"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C137"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D137"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA19"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B138"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B139"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C138"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D138"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C139"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D139"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA20"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B140"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B141"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C140"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D140"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C141"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D141"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA21"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B142"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B143"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C142"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D142"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C143"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D143"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA22"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B144"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B145"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C144"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D144"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C145"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D145"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA23"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B146"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B147"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C146"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D146"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C147"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D147"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA24"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B148"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B149"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C148"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D148"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C149"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D149"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'BPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "BPA25"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Bpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B150"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!B151"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C150"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D150"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!C151"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas BPA.xlsx]Sheet1'!D151"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
#########################################################################################CPA
    elif c == 'CPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA1"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B102"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B103"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C102"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D102"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C103"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D103"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA2"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B104"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B105"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C104"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D104"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C105"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D105"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA3"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B106"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B107"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C106"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D106"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C107"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D107"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA4"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B108"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B109"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C108"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D108"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C109"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D109"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA5"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B110"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B111"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C110"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C111"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D111"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA6"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B112"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B113"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C112"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D112"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C113"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D113"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA7"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B114"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B115"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C114"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D114"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C115"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D115"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA8"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B116"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B117"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C116"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D116"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C117"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D117"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA9"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B118"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B119"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C118"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D118"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C119"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D119"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA10"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B120"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B121"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C120"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C121"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D121"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA11"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B122"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B123"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C122"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D122"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C123"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D123"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA12"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B124"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B125"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C124"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D124"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C125"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D125"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA13"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B126"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B127"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C126"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D126"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C127"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D127"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA14"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B128"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B129"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C128"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D128"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C129"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D129"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA15"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B130"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B131"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C130"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D130"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C131"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D131"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA16"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B132"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B133"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C132"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D132"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C133"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D133"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA17"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B134"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B135"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C134"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D134"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C135"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D135"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA18"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B136"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B137"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C136"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D136"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C137"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D137"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA19"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B138"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B139"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C138"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D138"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C139"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D139"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA20"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B140"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B141"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C140"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D140"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C141"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D141"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA21"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B142"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B143"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C142"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D142"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C143"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D143"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA22"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B144"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B145"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C144"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D144"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C145"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D145"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA23"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B146"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B147"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C146"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D146"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C147"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D147"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA24"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B148"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B149"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C148"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D148"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C149"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D149"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'CPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "CPA25"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Cpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B150"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!B151"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C150"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D150"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!C151"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas CPA.xlsx]Sheet1'!D151"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
##############################################################################
    elif c == 'DPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA1"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B102"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B103"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C102"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D102"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C103"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D103"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA2"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B104"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B105"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C104"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D104"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C105"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D105"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA3"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B106"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B107"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C106"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D106"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C107"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D107"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA4"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B108"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B109"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C108"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D108"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C109"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D109"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA5"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B110"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B111"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C110"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C111"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D111"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA6"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B112"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B113"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C112"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D112"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C113"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D113"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA7"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B114"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B115"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C114"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D114"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C115"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D115"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA8"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B116"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B117"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C116"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D116"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C117"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D117"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA9"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B118"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B119"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C118"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D118"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C119"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D119"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA10"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B120"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B121"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C120"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C121"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D121"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA11"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B122"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B123"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C122"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D122"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C123"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D123"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA12"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B124"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B125"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C124"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D124"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C125"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D125"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA13"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B126"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B127"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C126"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D126"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C127"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D127"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA14"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B128"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B129"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C128"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D128"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C129"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D129"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA15"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B130"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B131"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C130"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D130"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C131"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D131"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA16"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B132"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B133"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C132"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D132"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C133"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D133"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA17"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B134"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B135"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C134"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D134"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C135"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D135"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA18"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B136"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B137"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C136"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D136"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C137"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D137"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA19"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B138"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B139"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C138"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D138"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C139"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D139"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    if c == 'DPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!A140"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!A141"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B140"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B141"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C140"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D140"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C141"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D141"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA21"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B142"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B143"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C142"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D142"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C143"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D143"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA22"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B144"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B145"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C144"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D144"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C145"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D145"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA23"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B146"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B147"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C146"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D146"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C147"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D147"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA24"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B148"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B149"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C148"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D148"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C149"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D149"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'DPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "DPA25"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Dpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B150"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!B151"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C150"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D150"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!C151"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas DPA.xlsx]Sheet1'!D151"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
###############################################################################
    elif c == 'EPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA1"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B102"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B103"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C102"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D102"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C103"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D103"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA2"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B104"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B105"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C104"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D104"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C105"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D105"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA3"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B106"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B107"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C106"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D106"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C107"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D107"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA4"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B108"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B109"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C108"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D108"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C109"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D109"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA5"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B110"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B111"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C110"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C111"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D111"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA6"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B112"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B113"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C112"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D112"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C113"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D113"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA7"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B114"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B115"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C114"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D114"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C115"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D115"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA8"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "EPa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B116"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B117"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C116"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D116"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C117"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D117"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA9"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B118"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B119"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C118"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D118"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C119"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D119"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA10"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B120"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B121"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C120"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C121"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D121"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'Epa11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA11"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B122"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B123"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C122"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D122"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C123"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D123"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA12"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B124"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B125"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C124"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D124"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C125"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D125"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA13"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B126"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B127"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C126"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D126"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C127"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D127"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA14"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B128"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B129"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C128"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D128"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C129"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D129"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA15"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B130"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B131"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C130"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D130"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C131"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D131"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA16"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B132"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B133"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C132"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D132"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C133"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D133"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA17"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B134"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B135"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C134"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D134"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C135"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D135"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA18"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B136"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B137"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C136"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D136"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C137"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D137"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA19"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B138"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B139"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C138"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D138"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C139"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D139"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    if c == 'EPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA20"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B140"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B141"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C140"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D140"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C141"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D141"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA21"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B142"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B143"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C142"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D142"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C143"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D143"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA22"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B144"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B145"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C144"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D144"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C145"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D145"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA23"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B146"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B147"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C146"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D146"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C147"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D147"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA24"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B148"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B149"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C148"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D148"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C149"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D149"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'EPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "EPA25"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Epa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B150"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!B151"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C150"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D150"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!C151"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas EPA.xlsx]Sheet1'!D151"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
##################################################################################
    elif c == 'FPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA1"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B102"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B103"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C102"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D102"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C103"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D103"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA2"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B104"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B105"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C104"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D104"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C105"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D105"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA3"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B106"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B107"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C106"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D106"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C107"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D107"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA4"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B108"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B109"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C108"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D108"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C109"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D109"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA5"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B110"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B111"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C110"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C111"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D111"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA6"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B112"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B113"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C112"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D112"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C113"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D113"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA7"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B114"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B115"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C114"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D114"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C115"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D115"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA8"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B116"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B117"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C116"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D116"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C117"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D117"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA9"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B118"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B119"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C118"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D118"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C119"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D119"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA10"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B120"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B121"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C120"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C121"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D121"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA11"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B122"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B123"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C122"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D122"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C123"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D123"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA12"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B124"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B125"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C124"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D124"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C125"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D125"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA13"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B126"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B127"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C126"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D126"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C127"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D127"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA14"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B128"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B129"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C128"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D128"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C129"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D129"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA15"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B130"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B131"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C130"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D130"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C131"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D131"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA16"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B132"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B133"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C132"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D132"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C133"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D133"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA17"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B134"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B135"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C134"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D134"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C135"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D135"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA18"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B136"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B137"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C136"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D136"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C137"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D137"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA19"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B138"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B139"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C138"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D138"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C139"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D139"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    if c == 'FPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!A140"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!A141"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B140"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B141"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C140"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D140"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C141"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D141"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA21"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B142"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B143"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C142"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D142"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C143"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D143"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA22"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B144"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B145"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C144"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D144"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C145"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D145"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA23"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B146"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B147"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C146"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D146"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C147"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D147"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA24"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B148"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B149"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C148"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D148"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C149"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D149"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'FPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "FPA25"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Fpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B150"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!B151"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C150"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D150"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!C151"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas FPA.xlsx]Sheet1'!D151"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
########################################################################################
    elif c == 'GPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA1"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B102"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B103"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C102"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D102"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C103"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D103"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA2"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B104"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B105"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C104"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D104"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C105"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D105"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA3"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B106"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B107"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C106"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D106"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C107"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D107"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA4"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B108"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B109"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C108"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D108"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C109"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D109"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA5"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B110"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B111"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C110"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C111"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D111"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA6"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B112"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B113"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C112"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D112"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C113"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D113"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA7"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B114"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B115"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C114"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D114"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C115"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D115"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA8"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B116"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B117"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C116"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D116"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C117"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D117"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA9"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B118"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B119"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C118"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D118"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C119"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D119"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA10"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B120"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B121"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C120"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C121"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D121"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA11"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B122"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B123"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C122"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D122"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C123"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D123"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA12"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B124"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B125"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C124"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D124"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C125"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D125"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA13"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B126"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B127"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C126"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D126"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C127"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D127"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA14"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B128"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B129"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C128"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D128"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C129"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D129"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA15"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B130"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B131"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C130"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D130"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C131"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D131"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA16"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B132"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B133"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C132"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D132"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C133"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D133"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA17"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B134"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B135"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C134"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D134"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C135"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D135"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA18"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B136"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B137"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C136"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D136"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C137"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D137"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA19"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B138"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B139"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C138"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D138"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C139"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D139"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    if c == 'GPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA20"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B140"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B141"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C140"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D140"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C141"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D141"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA21"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B142"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B143"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C142"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D142"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C143"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D143"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA22"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B144"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B145"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C144"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D144"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C145"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D145"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA23"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B146"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B147"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C146"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D146"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C147"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D147"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA24"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B148"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B149"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C148"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D148"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C149"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D149"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'GPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "GPA25"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Gpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B150"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!B151"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C150"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D150"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!C151"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas GPA.xlsx]Sheet1'!D151"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
#############################################################################
    elif c == 'HPA1':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA1"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B102"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B103"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C102"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D102"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C103"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D103"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA2':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA2"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B104"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B105"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C104"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D104"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C105"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D105"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA3':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA3"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B106"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B107"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C106"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D106"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C107"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D107"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA4':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA4"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B108"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B109"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C108"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D108"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C109"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D109"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA5':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA5"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B110"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B111"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C110"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C111"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D111"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA6':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA6"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B112"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B113"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C112"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D112"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C113"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D113"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA7':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA7"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B114"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B115"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C114"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D114"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C115"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D115"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA8':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA8"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B116"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B117"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C116"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D116"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C117"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D117"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA9':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA9"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B118"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B119"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C118"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D118"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C119"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D119"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA10':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA10"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B120"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B121"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C120"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D120"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C121"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D121"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA11':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA11"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B122"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B123"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C122"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D122"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C123"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D123"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA12':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA12"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B124"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B125"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C124"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D124"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C125"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D125"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA13':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA13"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B126"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B127"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C126"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D126"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C127"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D127"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA14':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA14"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B128"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B129"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C128"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D128"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C129"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D129"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA15':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA15"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B130"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B131"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C130"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D130"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C131"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D131"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA16':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA16"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B132"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B133"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C132"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D132"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C133"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D133"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA17':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA17"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B134"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B135"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C134"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D134"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C135"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D135"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA18':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA18"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B136"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B137"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C136"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D136"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C137"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D137"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA19':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA19"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B138"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B139"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C138"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D138"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C139"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D139"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    if c == 'HPA20':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA20"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B140"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B141"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C140"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D140"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C141"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D141"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA21':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA21"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B142"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B143"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C142"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D142"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C143"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D143"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA22':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA22"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B144"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B145"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C144"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D144"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C145"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D145"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA23':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA23"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B146"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B147"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C146"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D146"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C147"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D147"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA24':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA24"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B148"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B149"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C148"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D148"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C149"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D149"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")
    elif c == 'HPA25':
        wb = load_workbook("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        ws = wb.get_sheet_by_name("Sheet")
        kuning = ws.cell(row = A, column = B-1)
        kuning.value = "HPA25"
        kuning = ws.cell(row = A+1, column = B-1)
        kuning.value = "Hpa"
        kuning = ws.cell(row = A, column = B+2)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B150"
        kuning = ws.cell(row = A, column = B+3)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!B151"
        kuning = ws.cell(row = A, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C150"
        kuning = ws.cell(row = A, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D150"
        kuning = ws.cell(row = A+1, column = B)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!C151"
        kuning = ws.cell(row = A+1, column = F)
        kuning.value = "='/home/fakhril/PYTHON/BAGAN MAKER/BAGAN/PUTRA/[kelas HPA.xlsx]Sheet1'!D151"
        wb.save("/home/fakhril/PYTHON/BAGAN MAKER/JADWAL.xlsx")
        messagebox.showinfo("Success","Berhasil ditambahkan")

root.title("TAPAK SUCI TOURNAMENT ADMINISTRATOR")
canvas = Canvas(root,width=1000,height = 500)
canvas.pack()
label0 = Label(root,text="PENDAFTARAN PESERTA")
label0.config(font=("times",13))
canvas.create_window(230,50,window=label0)

label0 = Label(root,text="LIVE SCORE")
label0.config(font=("times",13))
canvas.create_window(700,50,window=label0)

label0 = Label(root,text="TRANSFER KE JADWAL")
label0.config(font=("times",13))
canvas.create_window(700,300,window=label0)

label0 = Label(root,text="Create : By Fakhril-AK")
label0.config(font=("times",13))
canvas.create_window(900,480,window=label0)
fno = str()
fname = str()
fkelas = str()
fkontingen = str()
##
label0 = Label(root,text="No")
label0.config(font=("helventica",10))
canvas.create_window(164,100,window=label0)

label0 = Label(root,text="Nama ")
label0.config(font=("helventica",10))
canvas.create_window(159,150,window=label0)

label0 = Label(root,text="Kelas ")
label0.config(font=("helventica",10))
canvas.create_window(158,200,window=label0)

label0 = Label(root,text="Kontingen ")
label0.config(font=("helventica",10))
canvas.create_window(143,250,window=label0)

label0 = Label(root,text="No Partai")
label0.config(font=("helventica",10))
canvas.create_window(580,100,window=label0)

label0 = Label(root,text="Kuning")
label0.config(font=("helventica",10))
canvas.create_window(584,150,window=label0)

label0 = Label(root,text="Biru ")
label0.config(font=("helventica",10))
canvas.create_window(594,200,window=label0)
##

ent = Entry(root)
canvas.create_window(260,100,window=ent)

ent1 = Entry(root)
canvas.create_window(260,150,window=ent1)

ent2 = Entry(root)
canvas.create_window(260,200,window=ent2)

ent3 = Entry(root)
canvas.create_window(260,250,window=ent3)

partai = Entry(root)
canvas.create_window(700, 100,window=partai)

kuning = Entry(root)
canvas.create_window(700,150,window=kuning)

biru = Entry(root)
canvas.create_window(700,200,window=biru)

read = Entry(root)
canvas.create_window(600,350,window=read)

write = Entry(root)
canvas.create_window(800,350,window=write)

buttonSub = Button(text='Enter', command=add, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 5)
canvas.create_window(120, 300, window=buttonSub)

buttonSub = Button(text='Delet', command=delet, bg='red', fg='white', font=('helvetica', 9, 'bold'), width = 5)
canvas.create_window(40, 300, window=buttonSub)

buttonSub = Button(text='SHOW ALL', command=show, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 8)
canvas.create_window(210, 300, window=buttonSub)

button = Button(text='CTRL', command=counter, bg='green', fg='white', font=('helvetica', 12, 'bold'), width = 15,height=3)
canvas.create_window(130, 400, window=button)

buttonSub = Button(text='KELAS A', command=classA, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 8)
canvas.create_window(300, 300, window=buttonSub)

buttonSub = Button(text='KELAS B', command=classB, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 8)
canvas.create_window(400, 300, window=buttonSub)

buttonSub = Button(text='KELAS C', command=classC, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 8)
canvas.create_window(300, 350, window=buttonSub)

buttonSub = Button(text='KELAS D', command=classD, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 8)
canvas.create_window(400, 350, window=buttonSub)

buttonSub = Button(text='KELAS E', command=classE, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 8)
canvas.create_window(300, 400, window=buttonSub)

buttonSub = Button(text='KELAS F', command=classF, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 8)
canvas.create_window(400, 400, window=buttonSub)

buttonSub = Button(text='KELAS G', command=classG, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 8)
canvas.create_window(300, 450, window=buttonSub)

buttonSub = Button(text='KELAS H', command=classH, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 8)
canvas.create_window(400, 450, window=buttonSub)

buttonSub = Button(text='live score', command=livescore, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 8)
canvas.create_window(700, 250, window=buttonSub)

buttonSub = Button(text='+ Jadwal', command=updatejadwal, bg='green', fg='white', font=('helvetica', 9, 'bold'), width = 8)
canvas.create_window(700, 400, window=buttonSub)
root.mainloop()
